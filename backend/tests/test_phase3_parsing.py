import io
from datetime import date

import pytest
from openpyxl import Workbook

from app.services.parsing import (
    CorruptedFileError,
    NoDataRowsError,
    UnsupportedFileTypeError,
    WrongColumnHeadersError,
    parse_upload_file,
)
from tests.test_phase3_uploads_api import CSV_HEADER

COLUMNS = CSV_HEADER.split(",")


def make_csv(*data_rows: dict[str, str]) -> bytes:
    """Builds a CSV with the full, portal-recognized header (CSV_HEADER) and
    one line per row dict, filling any column not given with an empty cell -
    lets each test only spell out the columns it actually cares about while
    still passing header validation."""
    lines = [CSV_HEADER]
    for row in data_rows:
        lines.append(",".join(row.get(col, "") for col in COLUMNS))
    return ("\n".join(lines) + "\n").encode()


def make_xlsx(header: list[str], *data_rows: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(header)
    for row in data_rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_csv_basic_rows():
    content = make_csv(
        {"NMC PIN": "16H0404E", "First Name": "ROSE 1", "Last Name": "LEE", "Institute Code": "1315"},
        {"NMC PIN": "16H0405E", "First Name": "ROSE 2", "Last Name": "LEE", "Institute Code": "1315"},
    )
    rows = parse_upload_file("students.csv", content)
    assert len(rows) == 2
    assert rows[0]["nmc_nmcpin"] == "16H0404E"
    assert rows[0]["nmc_firstname"] == "ROSE 1"
    # columns present in the header but blank in this row default to None
    assert rows[0]["nmc_dateofbirth"] is None


def test_parse_csv_skips_trailing_blank_lines():
    content = make_csv({"NMC PIN": "16H0404E", "First Name": "ROSE 1"}) + b"\n\n"
    rows = parse_upload_file("students.csv", content)
    assert len(rows) == 1


def test_parse_csv_trims_whitespace():
    content = make_csv({"NMC PIN": " 16H0404E ", "First Name": " ROSE 1 "})
    rows = parse_upload_file("students.csv", content)
    assert rows[0]["nmc_nmcpin"] == "16H0404E"
    assert rows[0]["nmc_firstname"] == "ROSE 1"


def test_parse_csv_ignores_previous_institute_code_column():
    # "Previous Institute Code" is a legitimate column per requirements.md's
    # file field mapping, but has no table field - its presence shouldn't
    # affect header validation or leak into the parsed row.
    extended_columns = COLUMNS + ["Previous Institute Code"]
    values = {"NMC PIN": "16H0404E", "Previous Institute Code": "9999"}
    header_line = ",".join(extended_columns)
    row_line = ",".join(values.get(col, "") for col in extended_columns)
    content = (header_line + "\n" + row_line + "\n").encode()

    rows = parse_upload_file("students.csv", content)
    assert rows[0]["nmc_nmcpin"] == "16H0404E"
    assert "nmc_previousinstitutecode" not in rows[0]
    assert "Previous Institute Code" not in rows[0]


def test_parse_csv_empty_content_is_corrupted():
    with pytest.raises(CorruptedFileError):
        parse_upload_file("students.csv", b"")


def test_parse_csv_wrong_column_headers_raises():
    content = "Employee ID,First,Last\n123,Rose,Lee\n".encode()
    with pytest.raises(WrongColumnHeadersError):
        parse_upload_file("students.csv", content)


def test_parse_csv_header_only_raises_no_data_rows():
    content = (CSV_HEADER + "\n").encode()
    with pytest.raises(NoDataRowsError):
        parse_upload_file("students.csv", content)


def test_parse_xlsx_basic_rows_and_date_conversion():
    header = COLUMNS
    row1 = [None] * len(COLUMNS)
    row1[COLUMNS.index("NMC PIN")] = "16H0404E"
    row1[COLUMNS.index("First Name")] = "ROSE 1"
    row1[COLUMNS.index("Last Name")] = "LEE"
    row1[COLUMNS.index("Date of Birth")] = date(2002, 5, 24)
    row2 = [None] * len(COLUMNS)
    row2[COLUMNS.index("NMC PIN")] = "16H0405E"
    row2[COLUMNS.index("Date of Birth")] = "20040321"
    content = make_xlsx(header, row1, row2)

    rows = parse_upload_file("students.xlsx", content)
    assert len(rows) == 2
    assert rows[0]["nmc_dateofbirth"] == "20020524"
    assert rows[1]["nmc_dateofbirth"] == "20040321"


def test_parse_xlsx_skips_fully_blank_rows():
    header = COLUMNS
    row1 = [None] * len(COLUMNS)
    row1[COLUMNS.index("NMC PIN")] = "16H0404E"
    blank_row = [None] * len(COLUMNS)
    row2 = [None] * len(COLUMNS)
    row2[COLUMNS.index("NMC PIN")] = "16H0405E"
    content = make_xlsx(header, row1, blank_row, row2)

    rows = parse_upload_file("students.xlsx", content)
    assert len(rows) == 2


def test_parse_xlsx_empty_content_is_corrupted():
    with pytest.raises(CorruptedFileError):
        parse_upload_file("students.xlsx", b"")


def test_parse_xlsx_sheet_with_no_rows_at_all_is_corrupted():
    wb = Workbook()
    ws = wb.active
    ws.delete_rows(1, ws.max_row)
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(CorruptedFileError):
        parse_upload_file("students.xlsx", buf.getvalue())


def test_parse_xlsx_wrong_column_headers_raises():
    content = make_xlsx(["Employee ID", "First", "Last"], ["123", "Rose", "Lee"])
    with pytest.raises(WrongColumnHeadersError):
        parse_upload_file("students.xlsx", content)


def test_parse_xlsx_header_only_raises_no_data_rows():
    content = make_xlsx(COLUMNS)
    with pytest.raises(NoDataRowsError):
        parse_upload_file("students.xlsx", content)


def test_unsupported_extension_raises():
    with pytest.raises(UnsupportedFileTypeError):
        parse_upload_file("students.txt", b"whatever")
