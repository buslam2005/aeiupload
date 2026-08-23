import csv
import io
from datetime import date, datetime
from zipfile import BadZipFile

from openpyxl import load_workbook

# Maps an uploaded file's own header names to upload_students' business
# columns, per requirements.md's "upload file field mapping" section - the
# file's headers are human-readable labels (e.g. "NMC PIN", "Course Code"),
# not the internal nmc_* names. "Previous Institute Code" has no table field
# and is intentionally omitted (dropped at parse time). A row missing a column
# here just gets None for it, which surfaces naturally as a mismatch (or, for
# nmc_nmcpin, as "record not found") during matching - no separate validation
# step needed at parse time.
FILE_COLUMN_TO_FIELD = {
    "NMC PIN": "nmc_nmcpin",
    "Title": "nmc_nmctitlename",
    "First Name": "nmc_firstname",
    "Middle Name": "nmc_maidenname",
    "Last Name": "nmc_lastname",
    "Date of Birth": "nmc_dateofbirth",
    "Gender": "nmc_gender",
    "Nationality": "nmc_nationalityname",
    "Place of Birth": "nmc_countryofbirthname",
    "Email Address": "nmc_email",
    "Address Line 1": "nmc_addressline1",
    "Address Line 2": "nmc_addressline2",
    "Address Line 3": "nmc_addressline3",
    "City": "nmc_city",
    "Postcode": "nmc_postcode",
    "Country": "nmc_countryname",
    "Institute Code": "nmc_traininginstitutecode",
    "Training Type": "nmc_trainingtype",
    "Course Code": "nmc_programme",
    "Academic Level": "nmc_academicroute",
    "Course Start Date": "nmc_coursestartdate",
    "Course End Date": "nmc_courseenddate",
    "Pass Date": "nmc_trainingexampassdate",
    "Start Date": "nmc_trainingstartdate",
    "End Date": "nmc_trainingcompletiondate",
}

# The columns the portal must recognize for a file to be processable at all -
# every key in FILE_COLUMN_TO_FIELD. "Previous Institute Code" is a legitimate
# column in the real template (per requirements.md) but isn't required here
# since it's dropped at parse time regardless of whether it's present.
REQUIRED_FILE_COLUMNS = set(FILE_COLUMN_TO_FIELD)


class UploadFileError(ValueError):
    """Base class for file-upload problems reported to the user as a 400,
    per requirements.md's "Error handling at file upload" section."""


class UnsupportedFileTypeError(UploadFileError):
    pass


class CorruptedFileError(UploadFileError):
    def __init__(self):
        super().__init__(
            "Portal fails to recognize the file. Please check the file before upload it again."
        )


class WrongColumnHeadersError(UploadFileError):
    def __init__(self):
        super().__init__("Column header(s) are wrong. Please check the file before upload it again.")


class NoDataRowsError(UploadFileError):
    def __init__(self):
        super().__init__(
            "There is no student record in the file. Please check the file before upload it again."
        )


def parse_upload_file(filename: str, content: bytes) -> list[dict[str, str | None]]:
    """Parse a .csv or .xlsx upload into a list of row dicts (one per data row,
    in file order, header/blank rows excluded), keyed by UPLOAD_COLUMNS.

    Raises a UploadFileError subclass - never lets a malformed file reach the
    caller as an unhandled exception - per requirements.md's "Error handling
    at file upload": a file the portal can't even read as a spreadsheet (no
    header, no rows, or corrupted) raises CorruptedFileError; a recognizable
    file whose header doesn't match the portal's expected columns raises
    WrongColumnHeadersError; a file with a valid header but zero data rows
    raises NoDataRowsError.
    """
    lower = filename.lower()
    if lower.endswith(".csv"):
        return _parse_csv(content)
    if lower.endswith(".xlsx"):
        return _parse_xlsx(content)
    raise UnsupportedFileTypeError(f"Unsupported file type: {filename}")


def _clean_cell(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y%m%d")
    text = str(value).strip()
    return text or None


def _rows_from_records(records: list[dict[str, object]]) -> list[dict[str, str | None]]:
    """Maps raw per-row dicts (keyed by the file's own header names) to
    upload_students fields, dropping rows that end up fully blank."""
    rows = []
    for raw_row in records:
        cleaned = {
            field: _clean_cell(raw_row.get(file_column))
            for file_column, field in FILE_COLUMN_TO_FIELD.items()
        }
        if any(cleaned.values()):
            rows.append(cleaned)
    return rows


def _parse_csv(content: bytes) -> list[dict[str, str | None]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise CorruptedFileError() from exc

    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise CorruptedFileError()
    if not REQUIRED_FILE_COLUMNS.issubset(reader.fieldnames):
        raise WrongColumnHeadersError()

    rows = _rows_from_records(list(reader))
    if not rows:
        raise NoDataRowsError()
    return rows


def _parse_xlsx(content: bytes) -> list[dict[str, str | None]]:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        header_row = next(rows_iter)
    except (BadZipFile, StopIteration, KeyError) as exc:
        raise CorruptedFileError() from exc

    header = [str(h).strip() if h is not None else "" for h in header_row]
    if not REQUIRED_FILE_COLUMNS.issubset(header):
        raise WrongColumnHeadersError()

    records = [dict(zip(header, raw)) for raw in rows_iter]
    rows = _rows_from_records(records)
    if not rows:
        raise NoDataRowsError()
    return rows
